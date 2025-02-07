from openpyxl import load_workbook

# 导入文件
wb = load_workbook(r'附件A 订购方案数据结果.xlsx')
# 选择工作表
sheet = wb['问题4的订购方案结果']
# 索引
sheet.cell(row=7, column=2).value = 6
sheet['A1'] = 'name'
# 保存文件
wb.save(r'附件A 订购方案数据结果.xlsx')